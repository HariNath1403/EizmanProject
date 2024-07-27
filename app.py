from flask import Flask, request, render_template, redirect, url_for, send_file
from werkzeug.utils import secure_filename
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Border, Alignment, Protection, NamedStyle, PatternFill
from openpyxl.cell.cell import MergedCell
from PIL import Image as PILImage
import io

app = Flask(__name__)
app.secret_key = 'supersecretkey'

# Folder to store uploaded files
# UPLOAD_FOLDER = 'uploads'
# app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# # Ensure the upload folder exists
# if not os.path.exists(UPLOAD_FOLDER):
#     os.makedirs(UPLOAD_FOLDER)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Get form data
        data_file = request.files['data-file']
        report_date = request.form['report-date']
        file_name = request.form['file-name']

        # Save the uploaded file
        # filename = secure_filename(data_file.filename)
        # file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        # data_file.save(file_path)

        if data_file:
            # Process the data without saving the file
            data_file_stream = io.BytesIO(data_file.read())
            members = read_process_data(data_file_stream)

        # Process the data
        # members = read_process_data(file_path)  # Update to return members

        # read_process_data(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        # output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{file_name}.xlsx')
        # generate_wb(members, output_path, report_date)

            output_stream = io.BytesIO()
            generate_wb(members, output_stream, report_date)   
            output_stream.seek(0) 

        # Send the generated file as a response
        # return send_file(output_path, as_attachment=True, download_name=f'{file_name}.xlsx')
    
        return send_file(output_stream, as_attachment=True, download_name=f'{file_name}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('index.html')

def copy_styles(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal,
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline,
            vertical=source_cell.border.vertical,
            horizontal=source_cell.border.horizontal
        )
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )
        target_cell.protection = Protection(
            locked=source_cell.protection.locked,
            hidden=source_cell.protection.hidden
        )
        target_cell.number_format = source_cell.number_format

def convert_to_numeric(value):
    # Handle missing values
    if pd.isnull(value):
        return np.nan

    if isinstance(value, (int, float)):
        return value
    
    # Remove currency symbols and commas
    value = str(value).replace('$', '').replace(',', '').strip()
    
    try:
        # Try to convert to a datetime object and then to a timestamp
        value = pd.to_datetime(value)
        return value.timestamp()
    except (ValueError, TypeError):
        try:
            # Attempt to convert to a numeric value
            return pd.to_numeric(value, errors='coerce')
        except (ValueError, TypeError):
            return np.nan
            
def check_number_instance(no):
    no = convert_to_numeric(no)
    if pd.notnull(no) and isinstance(no, (int, float)):
        return 1
    else:
        return 0

def format_excel(activesheet, destination):
    try: 
        for row in activesheet.iter_rows():
            for cell in row:
                new_cell = destination[cell.coordinate]
                new_cell.value = cell.value
                copy_styles(cell, new_cell)
                
        # Copy column widths
        for col in activesheet.column_dimensions:
            destination.column_dimensions[col].width = activesheet.column_dimensions[col].width

        # Copy row heights
        for row in activesheet.row_dimensions:
            destination.row_dimensions[row].height = activesheet.row_dimensions[row].height
        
        for row in activesheet.iter_rows():
            for cell in row:
                new_cell = destination[cell.coordinate]
                new_cell.value = cell.value
                copy_styles(cell, new_cell)
                
        # Copy merged cells
        for merged_cell in activesheet.merged_cells.ranges:
            destination.merge_cells(str(merged_cell))
        
        for row in activesheet.iter_rows():
            for cell in row:
                new_cell = destination[cell.coordinate]
                if not isinstance(cell, MergedCell):
                    new_cell.value = cell.value
                copy_styles(cell, new_cell)
        
        # Copy images
        # for image in activesheet._images:
        #     with open(image.ref, 'rb') as img_file:
        #         pil_img = PILImage.open(img_file)
        #         openpyxl_img = OpenpyxlImage(image.ref)
        #         destination.add_image(openpyxl_img, image.anchor)

        for image in activesheet._images:
            img = OpenpyxlImage(image.path)
            img.anchor = image.anchor
            destination.add_image(img)
    except Exception as e:
        print(f"Error copying styles or images: {e}")
    except:
        pass

def read_process_data(data_file):
    excel_file = pd.ExcelFile(data_file)

    report_last_no = 150
    columns = ["C", "M", "O", "Q", "S", "U", "W", "Y", "AA", "AC", "AE", "AG"]
    usecols = "C, M, O, Q, S, U, W, Y, AA, AC, AE, AG"

    df = {}
    last_no = 0
    last_row = 0

    for sheet_name in excel_file.sheet_names:
        frame = pd.read_excel(excel_file, sheet_name, header=None, usecols="C, M, O, Q, S, U, W, Y, AA, AC, AE, AG", skiprows=12, nrows=report_last_no,
            names=range(1, 13))  # Assuming 12 columns: C, M, O, Q, S, U, W, Y, AA, AC, AE, AG

        df[sheet_name] = frame
        last_no = sheet_name
        last_row = len(frame)

        if len(frame) < report_last_no:
            padding = pd.DataFrame(index=range(len(frame), 150), columns=columns)
            frame = pd.concat([frame, padding])

        df[sheet_name] = frame

    members = {}
    for x in range(0, last_row):
        try:
            curName = df[last_no].loc[x, 1]
            if pd.notna(curName):
                members[curName] = []
        except:
            pass

    for key in df.keys():
        # names = [df[last_no].loc[x, 1] for x in range(0, last_row, 10)]
        names = [x for x in members.keys()]
        mo_scores = []
        others_scores = []
        member_sum_mo = 0
        member_sum_oth = 0

        for x in range(0, last_row + 1):
            try:
                m_val = check_number_instance(df[key].loc[x, 2])
                o_val = check_number_instance(df[key].loc[x, 3])
                q_val = check_number_instance(df[key].loc[x, 4])
                s_val = check_number_instance(df[key].loc[x, 5])
                u_val = check_number_instance(df[key].loc[x, 6])
                w_val = check_number_instance(df[key].loc[x, 7])
                y_val = check_number_instance(df[key].loc[x, 8])
                aa_val = check_number_instance(df[key].loc[x, 9])
                ac_val = check_number_instance(df[key].loc[x, 10])
                ae_val = check_number_instance(df[key].loc[x, 11])
                ag_val = check_number_instance(df[key].loc[x, 12])
        
                member_sum_mo += m_val + o_val
                member_sum_oth += q_val + s_val + u_val + w_val + y_val + aa_val + ac_val + ae_val + ag_val
                
                if (x + 1) % 10 == 0 and x != 0:
                    mo_scores.append(member_sum_mo)
                    others_scores.append(member_sum_oth)          
                    member_sum_mo = 0
                    member_sum_oth = 0
            except:
                pass

        for i in range(0, len(names), 1):
            try:
                curName = names[i]
                if (pd.notna(curName)):
                    members[curName].append({key : {
                        'mo': mo_scores[i] * 1.0,
                        'others': others_scores[i] * 1.75
                        }
                    })
            except:
                pass

    return members

def generate_wb(fulldict, output_path, report_date):
    template_file = 'Template.xlsx'
    template_wb = load_workbook(template_file)
    template_sheet = template_wb['Template']
    
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    
    allNames = list(fulldict.keys())
        
    for name in allNames:
        new_sheet = new_wb.create_sheet(title=name)
        format_excel(template_sheet, new_sheet)
        
        values_mo = []
        values_others = []

        for x in range(0, 31):
            try:
                values_mo.append(fulldict[name][x]['{}'.format(x + 1)]['mo'])
                values_others.append(fulldict[name][x]['{}'.format(x + 1)]['others'])
            except:
                pass
            
        new_sheet['D5'] = name
        new_sheet['B50'] = name
        new_sheet['K6'] = report_date
        
        for x in range(0, 31):
            try:
                new_sheet['G{}'.format(11 + x)] = values_mo[x]
                new_sheet['I{}'.format(11 + x)] = values_mo[x]

                new_sheet['H{}'.format(11 + x)] = values_others[x]
                new_sheet['J{}'.format(11 + x)] = values_others[x] / 1.75
            except:
                pass
    
    new_wb.save(output_path)
    print('A new Excel document has been saved')

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
